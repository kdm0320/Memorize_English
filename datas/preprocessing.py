import string
import openpyxl

base_path = "/Users/gimdongmin/Desktop/memorize_english/datas/words"
file_name = "기본영어단어1000개.xlsx"
full_path = f"{base_path}/{file_name}"

new_file_name = f"new_{file_name}"
save_path = f"{base_path}/{new_file_name}"

half_index = 500
def from_two_colums_to_four():
    load_wb = openpyxl.load_workbook(full_path, data_only=True)
    load_ws = load_wb["Sheet1"]

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    first_words = load_ws["B"]
    second_words = load_ws["D"]

    for index, cell in enumerate(first_words):
        if index==0: continue
        word=""
        meaning=""
        for i in str(cell.value):
            if i.upper() != i.lower():
                word+=i
            else:
                meaning+=i
        new_ws[f"A{index+1}"] = index
        new_ws[f"B{index+1}"] = word
        new_ws[f"C{index+1}"] = meaning.lstrip()
            
                
    for index, cell in enumerate(second_words):
        if index==0: continue
        word=""
        meaning=""
        for i in str(cell.value):
            if i.upper() != i.lower():
                word+=i
            else:
                meaning+=i
        new_ws[f"D{index+1}"] = index+half_index
        new_ws[f"E{index+1}"] = word       
        new_ws[f"F{index+1}"] = meaning.lstrip()

    new_ws.merge_cells("A1:F1")
    new_ws["A1"] = "기본 영어단어 1000개"
  
    new_wb.save(save_path)


from_two_colums_to_four()
